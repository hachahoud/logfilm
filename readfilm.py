# read film names from excel file
from openpyxl import load_workbook

def get_films():
	films = []
	workbook = load_workbook(filename="myfilms.xlsx", data_only=True)
	for sheetname in workbook.sheetnames:
		# iterate through each sheet
		sheet = workbook[sheetname]
		for film in sheet.iter_rows(min_col=1,max_col=1,values_only=True):
			start = str(film).find("'") + 1
			end = str(film).rfind("'")
			film = str(film)[start:end]
			if "(None," in film: continue
			films.append(film)
	return films
