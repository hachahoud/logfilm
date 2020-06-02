from openpyxl import load_workbook, Workbook

# extract the films first
myfilms = {}
curr_year = ""
with open("list.txt","r") as myfile:
	for line in myfile:
		if line.strip().isnumeric():
			# add the year to highlight the beginning of a new sheet
			curr_year = line.strip()
			# make a new list for this year films
			myfilms[curr_year] = []

		else:
			# add film to the current year list
			# extract the film name
			film_name = ""
			# remove the '-' in the beginning of each name
			line = line[1:].strip()
			for letter in line:
				if letter in "  ',":
					# enable spaces in names
					film_name += letter
				elif not letter.isalnum():
					# name ends with a dot or other mark.
					break
				else:
					film_name += letter
			# add name to this year films
			myfilms[curr_year].append(film_name)

print(myfilms)
for year in myfilms.keys():
	print(myfilms[year])
	print(len(myfilms[year]))
input()

# open the spreadsheet
filename = "myfilms.xlsx"
workbook = Workbook()
# create a sheet for each year
for year in myfilms.keys():
	# the year is the title for the sheet
	workbook.create_sheet(year)
	#select current sheet
	sheet = workbook[year]
	# column fixe; first one
	# iteration through rows
	rows = len(myfilms[year])
	for row in range(rows):
		sheet.cell(row=row+1, column=1).value = myfilms[year][row]

workbook.save(filename=filename)
